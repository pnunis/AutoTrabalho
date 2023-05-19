import random
import openpyxl
import pandas as pd
import os

def criar_aposta():
    numeros = list(range(1, 26))
    aposta = random.sample(numeros, 15)
    return sorted(aposta)

def imprimir_aposta(aposta):
    print("Aposta gerada:")
    for numero in aposta:
        print(numero, end=" ")
    print()

def salvar_em_excel(apostas, diretorio):
    planilha = openpyxl.Workbook()
    planilha.create_sheet(title='Apostas', index=0)
    aba = planilha['Apostas']
    for i, aposta in enumerate(apostas, start=1):
        aba.cell(row=i, column=1, value=', '.join(str(num) for num in aposta))
    nome_arquivo = os.path.join(diretorio, 'apostas.xlsx')
    planilha.save(nome_arquivo)
    print(f"As apostas foram salvas no arquivo '{nome_arquivo}'.")

def salvar_em_txt(apostas, diretorio):
    nome_arquivo = os.path.join(diretorio, 'apostas.txt')
    with open(nome_arquivo, 'w') as arquivo:
        for aposta in apostas:
            arquivo.write(', '.join(str(num) for num in aposta) + '\n')
    print(f"As apostas foram salvas no arquivo '{nome_arquivo}'.")

# Exemplo de uso
if __name__ == '__main__':
    num_jogos = 8
    apostas = []
    for _ in range(num_jogos):
        aposta = criar_aposta()
        apostas.append(aposta)
        imprimir_aposta(aposta)
        print()
    diretorio = r'D:\Users\paulo.souza\Desktop\TESTE PYEXCEL\jogos'
    # Salvar as apostas em um arquivo Excel
    salvar_em_excel(apostas, diretorio)
    df = pd.read_excel(os.path.join(diretorio, 'apostas.xlsx'))
    # Salvar as apostas em um arquivo de texto
    salvar_em_txt(apostas, diretorio)
    df = pd.read_csv(os.path.join(diretorio, 'apostas.txt'))
